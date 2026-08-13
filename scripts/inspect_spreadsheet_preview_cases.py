#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def _marker_count(preview: str, marker: str) -> int:
    return sum(1 for line in preview.splitlines() if line.startswith(marker))


def _text(value: Any, limit: int = 160) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    return s if len(s) <= limit else s[:limit] + "..."


def _xlsx_states(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        return [(ws.title, ws.sheet_state == "visible") for ws in wb.worksheets]
    finally:
        wb.close()


def _xls_states(path: Path):
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=False)
    return [(ws.name, int(getattr(ws, "visibility", 0) or 0) == 0) for ws in wb.sheets()]


def _xlsx_rows(path: Path, sheet_names: list[str], max_nonempty_rows: int, max_cells: int):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    out = []
    try:
        for name in sheet_names:
            ws = wb[name]
            rows = []
            for r in range(1, int(ws.max_row or 0) + 1):
                cells = []
                for c in range(1, int(ws.max_column or 0) + 1):
                    v = ws.cell(r, c).value
                    t = _text(v)
                    if t:
                        cells.append({"col": c, "value": t})
                        if len(cells) >= max_cells:
                            break
                if cells:
                    rows.append({"row": r, "cells": cells})
                    if len(rows) >= max_nonempty_rows:
                        break
            out.append({
                "sheet": name,
                "visible": ws.sheet_state == "visible",
                "max_row": int(ws.max_row or 0),
                "max_col": int(ws.max_column or 0),
                "rows": rows,
            })
    finally:
        wb.close()
    return out


def _xls_rows(path: Path, sheet_names: list[str], max_nonempty_rows: int, max_cells: int):
    import xlrd
    wb = xlrd.open_workbook(path, formatting_info=False)
    out = []
    for name in sheet_names:
        ws = wb.sheet_by_name(name)
        rows = []
        for r in range(ws.nrows):
            cells = []
            for c in range(ws.ncols):
                t = _text(ws.cell_value(r, c))
                if t:
                    cells.append({"col": c + 1, "value": t})
                    if len(cells) >= max_cells:
                        break
            if cells:
                rows.append({"row": r + 1, "cells": cells})
                if len(rows) >= max_nonempty_rows:
                    break
        out.append({
            "sheet": name,
            "visible": int(getattr(ws, "visibility", 0) or 0) == 0,
            "max_row": int(ws.nrows or 0),
            "max_col": int(ws.ncols or 0),
            "rows": rows,
        })
    return out


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--implementation-root", required=True)
    p.add_argument("--fixture-root", required=True)
    p.add_argument("--output-dir", required=True)
    ns = p.parse_args()

    impl = Path(ns.implementation_root).resolve()
    fixtures = Path(ns.fixture_root).resolve()
    out_dir = Path(ns.output_dir).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    sys.path.insert(0, str(impl / "l0"))

    from l0_processor.config import Config
    from l0_processor.extractors import extract_local

    cfg = Config()
    records = []
    for path in sorted(fixtures.iterdir()):
        if not path.is_file() or path.suffix.lower() not in {".xls", ".xlsx", ".xlsm"}:
            continue
        preview, status, meta = extract_local(path, cfg)
        if status != "succeeded" or preview is None:
            continue
        states = _xls_states(path) if path.suffix.lower() == ".xls" else _xlsx_states(path)
        processed = {str(s.get("sheet")) for s in meta.get("sheets", [])}
        visible_names = [name for name, visible in states if visible]
        omitted_visible = [name for name in visible_names if name not in processed]
        blocks = sum(len(s.get("table_blocks") or []) for s in meta.get("sheets", []))
        headers = _marker_count(preview, "@@XLSX_TABLE_HEADER@@")
        rows = _marker_count(preview, "@@XLSX_TABLE_ROW@@")
        max_effective_col = max((int(s.get("effective_max_col") or 0) for s in meta.get("sheets", [])), default=0)
        records.append({
            "path": path,
            "preview": preview,
            "meta": meta,
            "states": states,
            "visible_names": visible_names,
            "omitted_visible": omitted_visible,
            "blocks": blocks,
            "headers": headers,
            "rows": rows,
            "max_effective_col": max_effective_col,
        })

    selected: list[tuple[str, dict[str, Any]]] = []
    used: set[str] = set()

    def pick(category: str, predicate):
        for rec in records:
            if rec["path"].name in used:
                continue
            if predicate(rec):
                selected.append((category, rec))
                used.add(rec["path"].name)
                return

    pick("table_detected_but_no_table_rows_in_preview", lambda r: r["blocks"] > 0 and r["rows"] == 0)
    pick("wide_table_with_header_omission", lambda r: r["blocks"] > r["headers"] and r["max_effective_col"] > cfg.table_preview.max_cols and len(r["visible_names"]) <= cfg.table_preview.max_sheets)
    pick("visible_sheets_omitted", lambda r: bool(r["omitted_visible"]))

    payload = {"cases": []}
    for category, rec in selected:
        path = rec["path"]
        if category == "visible_sheets_omitted":
            inspect_names = rec["visible_names"]
            max_rows = 18
            max_cells = 32
        elif category == "wide_table_with_header_omission":
            inspect_names = [name for name, _ in rec["states"] if name in {str(s.get('sheet')) for s in rec["meta"].get("sheets", [])}]
            max_rows = 30
            max_cells = 50
        else:
            inspect_names = [name for name, _ in rec["states"] if name in {str(s.get('sheet')) for s in rec["meta"].get("sheets", [])}]
            max_rows = 80
            max_cells = 50

        sheet_rows = _xls_rows(path, inspect_names, max_rows, max_cells) if path.suffix.lower() == ".xls" else _xlsx_rows(path, inspect_names, max_rows, max_cells)
        payload["cases"].append({
            "category": category,
            "file": path.name,
            "preview": rec["preview"],
            "preview_metadata": rec["meta"],
            "all_sheet_states": [{"sheet": n, "visible": v} for n, v in rec["states"]],
            "omitted_visible_sheets": rec["omitted_visible"],
            "actual_sheet_samples": sheet_rows,
        })

    (out_dir / "inspection.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"inspection_case_count={len(payload['cases'])}")
    print("inspection_categories=" + ",".join(c for c, _ in selected))
    print("SPREADSHEET_PRIVATE_INSPECTION_COMPLETE")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
