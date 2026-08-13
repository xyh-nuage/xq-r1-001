#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

MARKERS = (
    "@@XLSX_KV@@",
    "@@XLSX_TABLE_HEADER@@",
    "@@XLSX_TABLE_ROW@@",
    "@@XLSX_TEXT@@",
    "@@XLSX_MERGED@@",
)


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def marker_counts(preview: str) -> dict[str, int]:
    counts = {marker: 0 for marker in MARKERS}
    for line in preview.splitlines():
        for marker in MARKERS:
            if line.startswith(marker):
                counts[marker] += 1
                break
    return counts


def family_from_markers(counts: dict[str, int]) -> str:
    has_kv = counts["@@XLSX_KV@@"] > 0
    has_table = counts["@@XLSX_TABLE_HEADER@@"] > 0 or counts["@@XLSX_TABLE_ROW@@"] > 0
    has_text = counts["@@XLSX_TEXT@@"] > 0
    if has_kv and has_table and has_text:
        return "mixed_all"
    if has_kv and has_table:
        return "mixed_kv_table"
    if has_text and has_table:
        return "text_table"
    if has_text and has_kv:
        return "text_kv"
    if has_table:
        return "table_only"
    if has_kv:
        return "kv_only"
    if has_text:
        return "text_only"
    return "empty"


def _row_gap2_count(nonempty_cols: list[int]) -> int:
    return sum(1 for a, b in zip(nonempty_cols, nonempty_cols[1:]) if b - a == 2)


def xlsx_diagnostics(path: Path) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    formula_coords: dict[str, list[str]] = {}
    sheets: list[dict[str, Any]] = []
    totals = Counter()

    for ws in wb.worksheets:
        hidden_rows = sum(1 for dim in ws.row_dimensions.values() if dim.hidden)
        hidden_cols = sum(1 for dim in ws.column_dimensions.values() if dim.hidden)
        if ws.sheet_state != "visible":
            totals["hidden_sheets"] += 1

        nonempty_rows = 0
        max_nonempty_cells_in_row = 0
        gap2_rows = 0
        formula_count = 0
        percentage_count = 0
        coords: list[str] = []

        for row in ws.iter_rows():
            nonempty_cols: list[int] = []
            for cell in row:
                value = cell.value
                if value is not None and str(value).strip() != "":
                    nonempty_cols.append(cell.column)
                if cell.data_type == "f" or (
                    isinstance(value, str) and value.startswith("=")
                ):
                    formula_count += 1
                    coords.append(cell.coordinate)
                if value is not None and "%" in str(cell.number_format):
                    percentage_count += 1
            if nonempty_cols:
                nonempty_rows += 1
                max_nonempty_cells_in_row = max(
                    max_nonempty_cells_in_row, len(nonempty_cols)
                )
                if _row_gap2_count(sorted(nonempty_cols)):
                    gap2_rows += 1

        if coords:
            formula_coords[ws.title] = coords
        totals["formula_cells"] += formula_count
        totals["percentage_cells"] += percentage_count
        totals["hidden_rows"] += hidden_rows
        totals["hidden_cols"] += hidden_cols
        totals["merged_ranges"] += len(ws.merged_cells.ranges)
        totals["gap2_rows"] += gap2_rows

        sheets.append(
            {
                "sheet": ws.title,
                "state": ws.sheet_state,
                "max_row": ws.max_row or 0,
                "max_col": ws.max_column or 0,
                "nonempty_rows": nonempty_rows,
                "max_nonempty_cells_in_row": max_nonempty_cells_in_row,
                "merged_ranges": len(ws.merged_cells.ranges),
                "hidden_rows": hidden_rows,
                "hidden_cols": hidden_cols,
                "formula_cells": formula_count,
                "percentage_cells": percentage_count,
                "gap2_rows": gap2_rows,
            }
        )
    wb.close()

    missing_cached = 0
    if formula_coords:
        cached = openpyxl.load_workbook(path, read_only=False, data_only=True)
        try:
            for sheet_name, coords in formula_coords.items():
                ws = cached[sheet_name]
                for coord in coords:
                    value = ws[coord].value
                    if value is None or str(value).strip() == "":
                        missing_cached += 1
        finally:
            cached.close()

    return {
        "format": "xlsx",
        "formula_check_supported": True,
        "formula_missing_cached": missing_cached,
        "totals": dict(totals),
        "sheets": sheets,
    }


def xls_diagnostics(path: Path) -> dict[str, Any]:
    import xlrd

    wb = xlrd.open_workbook(path, formatting_info=False)
    sheets: list[dict[str, Any]] = []
    totals = Counter()
    for ws in wb.sheets():
        nonempty_rows = 0
        max_nonempty_cells_in_row = 0
        gap2_rows = 0
        for r0 in range(ws.nrows):
            nonempty_cols = [
                c0 + 1
                for c0 in range(ws.ncols)
                if ws.cell_type(r0, c0)
                not in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}
                and str(ws.cell_value(r0, c0)).strip() != ""
            ]
            if nonempty_cols:
                nonempty_rows += 1
                max_nonempty_cells_in_row = max(
                    max_nonempty_cells_in_row, len(nonempty_cols)
                )
                if _row_gap2_count(nonempty_cols):
                    gap2_rows += 1
        visibility = int(getattr(ws, "visibility", 0) or 0)
        if visibility:
            totals["hidden_sheets"] += 1
        merged_ranges = len(getattr(ws, "merged_cells", []) or [])
        totals["merged_ranges"] += merged_ranges
        totals["gap2_rows"] += gap2_rows
        sheets.append(
            {
                "sheet": ws.name,
                "visibility": visibility,
                "max_row": ws.nrows,
                "max_col": ws.ncols,
                "nonempty_rows": nonempty_rows,
                "max_nonempty_cells_in_row": max_nonempty_cells_in_row,
                "merged_ranges": merged_ranges,
                "hidden_rows": None,
                "hidden_cols": None,
                "formula_cells": None,
                "percentage_cells": None,
                "gap2_rows": gap2_rows,
            }
        )
    return {
        "format": "xls",
        "formula_check_supported": False,
        "formula_missing_cached": None,
        "totals": dict(totals),
        "sheets": sheets,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--implementation-root", required=True)
    parser.add_argument("--fixture-root", required=True)
    parser.add_argument("--implementation-sha", required=True)
    parser.add_argument("--fixture-sha", required=True)
    parser.add_argument("--output-dir", required=True)
    ns = parser.parse_args()

    implementation_root = Path(ns.implementation_root).resolve()
    fixture_root = Path(ns.fixture_root).resolve()
    output_dir = Path(ns.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    sys.path.insert(0, str(implementation_root / "l0"))
    from l0_processor.config import Config
    from l0_processor.extractors import extract_local

    cfg = Config()
    extensions = {".xls", ".xlsx", ".xlsm"}
    files = sorted(
        path
        for path in fixture_root.iterdir()
        if path.is_file() and path.suffix.lower() in extensions
    )
    if not files:
        print("SPREADSHEET_FIXTURE_SET_EMPTY")
        return 2

    details: list[dict[str, Any]] = []
    aggregate = Counter()
    family_counts = Counter()
    extension_counts = Counter()
    dataset_manifest = []

    for path in files:
        suffix = path.suffix.lower()
        extension_counts[suffix] += 1
        dataset_manifest.append(
            {
                "name": path.name,
                "size": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
        try:
            workbook = xls_diagnostics(path) if suffix == ".xls" else xlsx_diagnostics(path)
        except Exception as exc:
            workbook = {
                "format": suffix.lstrip("."),
                "diagnostic_error": type(exc).__name__,
                "formula_check_supported": suffix != ".xls",
                "formula_missing_cached": None,
                "totals": {},
                "sheets": [],
            }
            aggregate["workbook_diagnostic_failures"] += 1

        try:
            preview, status, meta = extract_local(path, cfg)
            extraction_error = None
        except Exception as exc:
            preview, status, meta = None, "exception", {}
            extraction_error = type(exc).__name__

        counts = marker_counts(preview or "")
        family = family_from_markers(counts)
        family_counts[family] += 1

        sheet_meta = meta.get("sheets", []) if isinstance(meta, dict) else []
        kv_pairs = [
            pair
            for sheet in sheet_meta
            for pair in (sheet.get("key_value_pairs") or [])
        ]
        table_blocks = [
            block
            for sheet in sheet_meta
            for block in (sheet.get("table_blocks") or [])
        ]
        inferred_pairs = sum(
            1 for pair in kv_pairs if pair.get("confidence") == "inferred_local"
        )
        strong_pairs = sum(
            1 for pair in kv_pairs if pair.get("confidence") == "strong"
        )
        multi_table_sheets = sum(
            1 for sheet in sheet_meta if len(sheet.get("table_blocks") or []) > 1
        )

        total_sheets = int(meta.get("total_sheets", 0) or 0) if isinstance(meta, dict) else 0
        sheets_processed = int(meta.get("sheets_processed", 0) or 0) if isinstance(meta, dict) else 0
        truncation_reasons: list[str] = []
        if total_sheets > sheets_processed:
            truncation_reasons.append("max_sheets")
        for sheet in workbook.get("sheets", []):
            if int(sheet.get("nonempty_rows") or 0) > int(cfg.table_preview.max_rows):
                truncation_reasons.append("max_rows")
                break
        for sheet in workbook.get("sheets", []):
            if int(sheet.get("max_nonempty_cells_in_row") or 0) > int(cfg.table_preview.max_cols):
                truncation_reasons.append("max_cols")
                break
        truncation_reasons = sorted(set(truncation_reasons))

        flags: list[str] = []
        if status != "succeeded":
            flags.append("extraction_failed")
            aggregate["extraction_failures"] += 1
        else:
            aggregate["extraction_success"] += 1
        if inferred_pairs:
            flags.append("inferred_kv")
            aggregate["inferred_kv_files"] += 1
        if multi_table_sheets:
            flags.append("multi_table")
            aggregate["multi_table_files"] += 1
        if truncation_reasons:
            flags.append("truncated")
            aggregate["truncated_files"] += 1

        totals = workbook.get("totals", {})
        if int(totals.get("hidden_sheets", 0) or 0):
            flags.append("hidden_sheet")
            aggregate["hidden_sheet_files"] += 1
        if int(totals.get("hidden_rows", 0) or 0):
            flags.append("hidden_rows")
            aggregate["hidden_row_files"] += 1
        if int(totals.get("hidden_cols", 0) or 0):
            flags.append("hidden_cols")
            aggregate["hidden_col_files"] += 1
        if int(totals.get("merged_ranges", 0) or 0):
            aggregate["merged_range_files"] += 1
        if int(totals.get("gap2_rows", 0) or 0):
            flags.append("gap2_structure_risk")
            aggregate["gap2_risk_files"] += 1
        if int(totals.get("formula_cells", 0) or 0):
            flags.append("formula_cells")
            aggregate["formula_files"] += 1
        missing_cached = workbook.get("formula_missing_cached")
        if isinstance(missing_cached, int) and missing_cached > 0:
            flags.append("formula_missing_cached")
            aggregate["formula_missing_cached_files"] += 1
        if int(totals.get("percentage_cells", 0) or 0):
            flags.append("percentage_display_risk")
            aggregate["percentage_format_files"] += 1
        if family in {"mixed_kv_table", "mixed_all"}:
            aggregate["mixed_kv_table_files"] += 1
        if family == "text_only":
            aggregate["text_only_files"] += 1
        if family == "empty":
            aggregate["empty_preview_files"] += 1

        aggregate["kv_pairs"] += len(kv_pairs)
        aggregate["kv_pairs_strong"] += strong_pairs
        aggregate["kv_pairs_inferred"] += inferred_pairs
        aggregate["table_blocks"] += len(table_blocks)
        aggregate["formula_cells"] += int(totals.get("formula_cells", 0) or 0)
        aggregate["formula_missing_cached_cells"] += (
            missing_cached if isinstance(missing_cached, int) else 0
        )
        aggregate["percentage_cells"] += int(totals.get("percentage_cells", 0) or 0)
        aggregate["hidden_sheets"] += int(totals.get("hidden_sheets", 0) or 0)
        aggregate["hidden_rows"] += int(totals.get("hidden_rows", 0) or 0)
        aggregate["hidden_cols"] += int(totals.get("hidden_cols", 0) or 0)
        aggregate["merged_ranges"] += int(totals.get("merged_ranges", 0) or 0)
        aggregate["gap2_rows"] += int(totals.get("gap2_rows", 0) or 0)

        details.append(
            {
                "file": path.name,
                "extension": suffix,
                "bytes": path.stat().st_size,
                "sha256": dataset_manifest[-1]["sha256"],
                "extractor": {
                    "status": status,
                    "error_type": extraction_error,
                    "preview_format": meta.get("preview_format") if isinstance(meta, dict) else None,
                    "sheets_processed": sheets_processed,
                    "total_sheets": total_sheets,
                    "marker_counts": counts,
                    "family": family,
                    "kv_pairs": len(kv_pairs),
                    "kv_pairs_strong": strong_pairs,
                    "kv_pairs_inferred": inferred_pairs,
                    "table_blocks": len(table_blocks),
                    "multi_table_sheets": multi_table_sheets,
                    "truncation_reasons": truncation_reasons,
                },
                "workbook": workbook,
                "flags": sorted(flags),
            }
        )

    manifest_payload = json.dumps(
        dataset_manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    fixture_set_sha256 = hashlib.sha256(manifest_payload).hexdigest()

    summary = {
        "schema_version": "spreadsheet_fixture_audit.v1",
        "implementation_sha": ns.implementation_sha,
        "fixture_sha": ns.fixture_sha,
        "fixture_set_sha256": fixture_set_sha256,
        "config": {
            "max_sheets": int(cfg.table_preview.max_sheets),
            "max_rows": int(cfg.table_preview.max_rows),
            "max_cols": int(cfg.table_preview.max_cols),
        },
        "files_total": len(files),
        "extension_counts": dict(sorted(extension_counts.items())),
        "family_counts": dict(sorted(family_counts.items())),
        "aggregate": dict(sorted(aggregate.items())),
        "no_llm": True,
        "semantic_changes": False,
    }
    details_payload = {
        "schema_version": "spreadsheet_fixture_audit_details.v1",
        "implementation_sha": ns.implementation_sha,
        "fixture_sha": ns.fixture_sha,
        "fixture_set_sha256": fixture_set_sha256,
        "files": details,
    }

    summary_path = output_dir / "summary.json"
    details_path = output_dir / "details.json"
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    details_path.write_text(
        json.dumps(details_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )

    print(f"fixture_files_total={len(files)}")
    for ext, count in sorted(extension_counts.items()):
        print(f"extension_{ext.lstrip('.')}={count}")
    for key in (
        "extraction_success",
        "extraction_failures",
        "mixed_kv_table_files",
        "inferred_kv_files",
        "kv_pairs",
        "kv_pairs_strong",
        "kv_pairs_inferred",
        "table_blocks",
        "multi_table_files",
        "truncated_files",
        "formula_files",
        "formula_cells",
        "formula_missing_cached_files",
        "formula_missing_cached_cells",
        "percentage_format_files",
        "percentage_cells",
        "hidden_sheet_files",
        "hidden_row_files",
        "hidden_col_files",
        "merged_range_files",
        "gap2_risk_files",
        "gap2_rows",
        "text_only_files",
        "empty_preview_files",
        "workbook_diagnostic_failures",
    ):
        print(f"{key}={int(aggregate.get(key, 0))}")
    print(f"fixture_set_sha256={fixture_set_sha256}")
    print(f"summary_sha256={sha256_file(summary_path)}")
    print(f"details_sha256={sha256_file(details_path)}")
    print("SPREADSHEET_FIXTURE_AUDIT_COMPLETE")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
